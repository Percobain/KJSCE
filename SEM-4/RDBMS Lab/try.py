import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation  # Add this import

def create_horizontal_database():
    wb = Workbook()
    ws = wb.active
    ws.title = "FinalOneONGOD"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    pk_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    fk_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Entity definitions with their attributes
    tables = {
        "DOCTOR": [
            ("doctor_id", "PK"),
            ("name", ""),
            ("contact_no", ""),
            ("DOB", ""),
            ("email", ""),
            ("department_id", "FK")
        ],
        "DEPARTMENT": [
            ("department_id", "PK"),
            ("name", ""),
            ("contact_no", ""),
            ("head_doctor", "")
        ],
        "PATIENT": [
            ("patient_id", "PK"),
            ("name", ""),
            ("DOB", ""),
            ("contact_no", ""),
            ("blood_group", ""),
            ("bed_id", "FK")
        ],
        "HOSPITAL_BED": [
            ("bed_id", "PK"),
            ("type", ""),
            ("status", ""),
            ("room_no", ""),
            ("ward_no", "")
        ],
        "ADDRESS": [
            ("address_id", "PK"),
            ("patient_id", "FK"),
            ("flat_no", ""),
            ("area", ""),
            ("pincode", ""),
            ("state", ""),
            ("city", ""),
            ("landmark", "")
        ],
        "APPOINTMENT": [
            ("appointment_id", "PK"),
            ("patient_id", "FK"),
            ("doctor_id", "FK"),
            ("date", ""),
            ("status", ""),
            ("fees", "")
        ],
        "MEDICAL_RECORD": [
            ("record_id", "PK"),
            ("patient_id", "FK"),
            ("date", ""),
            ("diagnosis", "")
        ],
        "BILL": [
            ("bill_id", "PK"),
            ("appointment_id", "FK"),
            ("date", ""),
            ("amount", ""),
            ("status", "")
        ],
        "PRESCRIPTION": [
            ("prescription_id", "PK"),
            ("appointment_id", "FK"),
            ("date", ""),
            ("medicine", ""),
            ("dosage", "")
        ]
    }
    
    sample_data = {
        "DOCTOR": [
            (1, "Dr. John Smith", "9876543210", "1980-05-15", "john.smith@hospital.com", 1),
            (2, "Dr. Sarah Brown", "9876543211", "1985-08-22", "sarah.brown@hospital.com", 2)
        ],
        "DEPARTMENT": [
            (1, "Cardiology", "9876543001", "Dr. John Smith"),
            (2, "Pediatrics", "9876543002", "Dr. Sarah Brown")
        ],
        "PATIENT": [
            (1, "James Wilson", "1990-03-25", "9876540001", "O+", 1),
            (2, "Emma Davis", "1988-07-12", "9876540002", "A+", 2)
        ],
        "HOSPITAL_BED": [
            (1, "Private", "Occupied", "101", "A"),
            (2, "General", "Available", "102", "B")
        ],
        "ADDRESS": [
            (1, 1, "Flat 101", "Downtown", "400001", "Maharashtra", "Mumbai", "Near Park"),
            (2, 2, "Flat 202", "Uptown", "400002", "Maharashtra", "Mumbai", "Near Mall")
        ],
        "APPOINTMENT": [
            (1, 1, 1, "2024-03-15", "Scheduled", 500),
            (2, 2, 2, "2024-03-16", "Completed", 600)
        ],
        "MEDICAL_RECORD": [
            (1, 1, "2024-03-15", "Fever and cold"),
            (2, 2, "2024-03-16", "Regular checkup")
        ],
        "BILL": [
            (1, 1, "2024-03-15", 500, "Pending"),
            (2, 2, "2024-03-16", 600, "Paid")
        ],
        "PRESCRIPTION": [
            (1, 1, "2024-03-15", "Paracetamol", "1-0-1"),
            (2, 2, "2024-03-16", "Vitamins", "0-1-0")
        ]
    }


    current_col = 1
    space_between_tables = 2

    for table_name, columns in tables.items():
        # Write table name
        cell = ws.cell(row=1, column=current_col, value=table_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        ws.merge_cells(start_row=1, start_column=current_col, 
                      end_row=1, end_column=current_col + len(columns) - 1)
        
        # Write column headers
        for col_idx, (col_name, key_type) in enumerate(columns):
            cell = ws.cell(row=2, column=current_col + col_idx, value=col_name)
            cell.border = border
            cell.font = Font(bold=True)
            
            # Color coding for PK/FK
            if (key_type == "PK"):
                cell.fill = pk_fill
            elif (key_type == "FK"):
                cell.fill = fk_fill
        
        # Add data validation for specific columns
        if table_name == "PATIENT":
            blood_group_col = current_col + 4  # blood_group column
            for row in range(3, 53):  # 50 rows for data
                cell = ws.cell(row=row, column=blood_group_col)
                dv = DataValidation(type="list", formula1='"A+,A-,B+,B-,AB+,AB-,O+,O-"')
                ws.add_data_validation(dv)
                dv.add(cell)

        if table_name == "HOSPITAL_BED":
            status_col = current_col + 2  # status column
            for row in range(3, 53):
                cell = ws.cell(row=row, column=status_col)
                dv = DataValidation(type="list", formula1='"Available,Occupied"')
                ws.add_data_validation(dv)
                dv.add(cell)

        if table_name == "APPOINTMENT":
            status_col = current_col + 4  # status column
            for row in range(3, 53):
                cell = ws.cell(row=row, column=status_col)
                dv = DataValidation(type="list", formula1='"Scheduled,Completed,Cancelled"')
                ws.add_data_validation(dv)
                dv.add(cell)

        # Add empty cells for data entry (50 rows)
        for row in range(3, 53):
            for col_idx in range(len(columns)):
                cell = ws.cell(row=row, column=current_col + col_idx)
                cell.border = Border(
                    left=Side(style='dotted'),
                    right=Side(style='dotted'),
                    top=Side(style='dotted'),
                    bottom=Side(style='dotted')
                )
        
        # Add sample data
        if table_name in sample_data:
            for row_idx, row_data in enumerate(sample_data[table_name], start=3):
                for col_idx, value in enumerate(row_data):
                    cell = ws.cell(row=row_idx, column=current_col + col_idx, value=value)

        # Set column widths
        for i in range(len(columns)):
            col_letter = get_column_letter(current_col + i)
            ws.column_dimensions[col_letter].width = 15

        # Move to next table position
        current_col += len(columns) + space_between_tables

    # Add relationships at the bottom
    relationship_row = 54
    relationships = [
        "RELATIONSHIPS:",
        "1. DOCTOR (1) ---- (N) APPOINTMENT",
        "2. PATIENT (1) ---- (N) APPOINTMENT",
        "3. DEPARTMENT (1) ---- (N) DOCTOR",
        "4. PATIENT (1) ---- (1) HOSPITAL_BED",
        "5. PATIENT (1) ---- (N) MEDICAL_RECORD",
        "6. PATIENT (1) ---- (1) ADDRESS",
        "7. APPOINTMENT (1) ---- (1) BILL",
        "8. APPOINTMENT (1) ---- (1) PRESCRIPTION"
    ]

    for i, rel in enumerate(relationships):
        cell = ws.cell(row=relationship_row + i, column=1, value=rel)
        cell.font = Font(bold=True if i == 0 else False)

    wb.save('final.xlsx')

if __name__ == "__main__":
    create_horizontal_database()